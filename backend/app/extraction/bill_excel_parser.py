"""Parses a general-purpose bill Excel (one row per invoice or multi-row per invoice)
for either Sales or Purchase — for the case where a user has their bills already
listed in a spreadsheet instead of (or alongside) photos. Column names vary
by user/supplier, so this scores header rows against known aliases rather
than assuming a fixed template.

Multi-line item rows or slab-wise breakdown columns sharing the same invoice number
and party are automatically consolidated into a single invoice object containing
a `rate_breakdown` list for Tally voucher generation.
"""
import re
from dataclasses import dataclass, field

import openpyxl
import xlrd


class _XlrdSheetAdapter:
    """Thin wrapper so the rest of this module can call .cell(row, col)
    with 1-based indices and .value, matching the openpyxl API we already
    use, regardless of whether the file is legacy .xls (xlrd) or .xlsx
    (openpyxl)."""

    class _Cell:
        __slots__ = ("value",)

        def __init__(self, value):
            self.value = value

    def __init__(self, xlrd_sheet):
        self._sheet = xlrd_sheet
        self.max_row = xlrd_sheet.nrows
        self.max_column = xlrd_sheet.ncols

    def cell(self, row, col):
        if row < 1 or col < 1 or row > self.max_row or col > self.max_column:
            return self._Cell(None)
        value = self._sheet.cell_value(row - 1, col - 1)
        return self._Cell(value if value != "" else None)


def _load_first_sheet(file_path: str):
    if file_path.lower().endswith(".xls"):
        wb = xlrd.open_workbook(file_path)
        return _XlrdSheetAdapter(wb.sheet_by_index(0))
    wb = openpyxl.load_workbook(file_path, data_only=True)
    return wb[wb.sheetnames[0]]


def _normalize(text) -> str:
    if text is None:
        return ""
    return re.sub(r"[₹()%\s\-_/.*?]", "", str(text)).lower()


def _number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[₹,\s]", "", str(value))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:
        return None


HEADER_VARIANTS = {
    "party": ["party", "partyname", "customername", "customer", "suppliername", "supplier", "name", "buyername", "sri", "vendorname", "vendor"],
    "date": ["date", "billdate", "invoicedate", "trandate"],
    "invoice_number": ["invoiceno", "invoicenumber", "billno", "billnumber", "invno", "tranno"],
    "taxable_value": ["taxablevalue", "taxableamount", "amount", "netamount", "value"],
    "gst_rate": ["gstrate", "taxrate", "rate", "gst%"],
    "cgst_amount": ["cgstamount", "cgst", "cgstamt"],
    "sgst_amount": ["sgstamount", "sgst", "sgstutamount", "sgstamt"],
    "igst_amount": ["igstamount", "igst", "igstamt"],
    "total_value": ["totalvalue", "totalamount", "grandtotal", "total", "billamount", "invoiceamount"],
}

STANDARD_GST_RATES = [0, 0.25, 1, 1.5, 3, 5, 6, 7.5, 12, 13.8, 18, 28]

_RATE_BUCKET_RE = re.compile(r"(value|cgst|sgst|igst)\s*@\s*([\d.]+)\s*%?", re.IGNORECASE)


def _closest_rate(pct):
    if pct is None:
        return 0.0
    nearest = min(STANDARD_GST_RATES, key=lambda r: abs(r - pct))
    return nearest if abs(nearest - pct) <= 0.5 else round(pct, 2)


def _find_rate_buckets(sheet, header_row):
    buckets: dict[float, dict] = {}
    for col in range(1, sheet.max_column + 1):
        header_text = str(sheet.cell(header_row, col).value or "").strip()
        m = _RATE_BUCKET_RE.fullmatch(header_text)
        if not m:
            continue
        kind, quoted_rate = m.group(1).lower(), float(m.group(2))
        rate = quoted_rate * 2 if kind in ("cgst", "sgst") else quoted_rate
        bucket = buckets.setdefault(rate, {"value_col": None, "cgst_col": None, "sgst_col": None, "igst_col": None})
        bucket[f"{kind}_col"] = col
    return [{"rate": rate, **cols} for rate, cols in sorted(buckets.items()) if cols["value_col"]]


def _find_header_row(sheet):
    best_row, best_cols, best_score = None, {}, 0
    for row in range(1, min(sheet.max_row, 15) + 1):
        found = {}
        for col in range(1, sheet.max_column + 1):
            value = _normalize(sheet.cell(row, col).value)
            if not value:
                continue
            for field_name, variants in HEADER_VARIANTS.items():
                if field_name not in found and value in variants:
                    found[field_name] = col
        has_amount = "taxable_value" in found or "total_value" in found
        score = len(found)
        if has_amount and score > best_score:
            best_row, best_cols, best_score = row, found, score
    if best_row is None:
        return None
    return best_row, best_cols


@dataclass
class ParsedBillRow:
    party: str | None
    date: str | None
    invoice_number: str | None
    taxable_value: float
    gst_rate: float
    cgst: float
    sgst: float
    igst: float
    total_value: float
    warnings: list = field(default_factory=list)
    split_from_multi_rate: bool = False
    rate_breakdown: list = field(default_factory=list)
    # [{"name": ..., "hsn": ..., "qty": ..., "unit": ..., "price": ..., "amount": ..., "rate": ...}, ...]
    # Populated only by the item-wise Tally-template parser below.
    items: list = field(default_factory=list)
    vch_type: str | None = None


def _excel_serial_to_date(serial: float):
    from datetime import datetime, timedelta
    try:
        return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    except (OverflowError, OSError, ValueError):
        return None


_DMY_RE = re.compile(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$")
_YMD_RE = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$")


def _parse_date_cell(date_val) -> str | None:
    """Normalizes a date cell (Excel datetime, serial number, or text) to
    an ISO 'YYYY-MM-DD' string. Text dates are assumed day-first
    (DD-MM-YYYY / DD/MM/YYYY) since that's the standard Indian/Tally
    convention this template is exported in — datetime.fromisoformat()
    downstream would otherwise silently misparse (or outright reject,
    raising MissingVoucherDateError at push time) a value like
    '01-04-2026' because it isn't already ISO-ordered.
    """
    if date_val is None:
        return None
    if hasattr(date_val, "date") and not isinstance(date_val, str):
        try:
            return str(date_val.date())
        except (TypeError, ValueError):
            pass
    if isinstance(date_val, (int, float)) and date_val > 1000:
        parsed = _excel_serial_to_date(date_val)
        return str(parsed) if parsed else None

    text = str(date_val).strip()
    if not text:
        return None

    m = _YMD_RE.match(text)
    if m:
        y, mo, d = m.groups()
        try:
            from datetime import date
            return str(date(int(y), int(mo), int(d)))
        except ValueError:
            return None

    m = _DMY_RE.match(text)
    if m:
        d, mo, y = m.groups()
        try:
            from datetime import date
            return str(date(int(y), int(mo), int(d)))
        except ValueError:
            return None

    return text


def _consolidate_invoice_rows(raw_rows: list[ParsedBillRow]) -> list[ParsedBillRow]:
    """Groups multiple Excel rows sharing the same party and invoice_number into
    a single invoice object with an aggregated rate_breakdown list.
    """
    grouped_map: dict[tuple[str, str], ParsedBillRow] = {}
    ungrouped: list[ParsedBillRow] = []

    for r in raw_rows:
        inv_no = (r.invoice_number or "").strip().lower()
        party_name = (r.party or "").strip().lower()

        line_breakdown = r.rate_breakdown if r.rate_breakdown else [{
            "rate": r.gst_rate,
            "taxable_value": r.taxable_value,
            "cgst": r.cgst,
            "sgst": r.sgst,
            "igst": r.igst
        }]

        if not inv_no:
            r.rate_breakdown = line_breakdown
            ungrouped.append(r)
            continue

        key = (party_name, inv_no)
        if key not in grouped_map:
            grouped_map[key] = ParsedBillRow(
                party=r.party,
                date=r.date,
                invoice_number=r.invoice_number,
                taxable_value=r.taxable_value,
                gst_rate=r.gst_rate,
                cgst=r.cgst,
                sgst=r.sgst,
                igst=r.igst,
                total_value=r.total_value,
                warnings=list(r.warnings),
                split_from_multi_rate=r.split_from_multi_rate,
                rate_breakdown=list(line_breakdown),
                items=list(r.items),
                vch_type=r.vch_type,
            )
        else:
            existing = grouped_map[key]
            existing.taxable_value = round(existing.taxable_value + r.taxable_value, 2)
            existing.cgst = round(existing.cgst + r.cgst, 2)
            existing.sgst = round(existing.sgst + r.sgst, 2)
            existing.igst = round(existing.igst + r.igst, 2)
            existing.total_value = round(existing.total_value + r.total_value, 2)
            existing.rate_breakdown.extend(line_breakdown)
            existing.items.extend(r.items)
            # Prefer the date of the row that actually has one — a
            # continuation item row in the template can come in blank
            # (Tally only requires the date on the invoice, not every line).
            if not existing.date and r.date:
                existing.date = r.date

            for w in r.warnings:
                if w not in existing.warnings:
                    existing.warnings.append(w)

            if r.split_from_multi_rate:
                existing.split_from_multi_rate = True

    consolidated = list(grouped_map.values()) + ungrouped

    # A "no date" warning raised by one line of a multi-line invoice is
    # stale once another line of the *same* invoice supplied the date used
    # above — don't leave it in the merged warnings list to confuse the
    # approver.
    for row in consolidated:
        if row.date:
            row.warnings = [w for w in row.warnings if "No date found" not in w]

    # Merge rate_breakdown entries with the same GST rate inside each invoice
    for row in consolidated:
        merged_breakdown: dict[float, dict] = {}
        for b in row.rate_breakdown:
            rt = float(b["rate"])
            if rt not in merged_breakdown:
                merged_breakdown[rt] = {
                    "rate": rt,
                    "taxable_value": float(b["taxable_value"]),
                    "cgst": float(b["cgst"]),
                    "sgst": float(b["sgst"]),
                    "igst": float(b["igst"])
                }
            else:
                m = merged_breakdown[rt]
                m["taxable_value"] = round(m["taxable_value"] + float(b["taxable_value"]), 2)
                m["cgst"] = round(m["cgst"] + float(b["cgst"]), 2)
                m["sgst"] = round(m["sgst"] + float(b["sgst"]), 2)
                m["igst"] = round(m["igst"] + float(b["igst"]), 2)

        row.rate_breakdown = list(merged_breakdown.values())
        if len(row.rate_breakdown) == 1:
            row.gst_rate = row.rate_breakdown[0]["rate"]

    return consolidated


# --- Item-wise ("Tally voucher import") template ------------------------
# Tally Prime's own "Sales Voucher" / "Purchase Voucher" Excel export/import
# template: one row per stock-item line (several rows can share the same
# Vch. No. when an invoice has more than one item), with a fixed set of
# per-slab GST columns ("IGST 5%", "CGST 2.5%", "SGST 2.5%", ...) rather
# than the freeform "gst_rate" + "cgst_amount" columns the generic parser
# above expects. Detected up front so a user who exports straight from
# Tally (or fills this exact template) doesn't hit the generic parser's
# "could not find a recognizable bill table" error.

_ITEM_TEMPLATE_HEADERS = {
    "date": ["vchdate"],
    "vch_type": ["vchtype"],
    "invoice_number": ["vchno"],
    "party": ["partyledger"],
    "sales_ledger": ["salesledger"],
    "stock_item": ["stockitem"],
    "quantity": ["quantity"],
    "rate": ["rate"],
    "unit": ["unit"],
    "amount": ["amount"],
    "narration": ["narration"],
    "gst_rate": ["gstrate"],
    "hsn": ["hsn"],
    "round_off": ["roundoff"],
}

# (gst_rate, igst_header_key, cgst_header_key, sgst_header_key) — header
# keys are post-_normalize (spaces/%/./₹ stripped), matching how "IGST 5%"
# and "CGST 2.5%" come out of _normalize().
_ITEM_TEMPLATE_RATE_SLABS = [
    (5, "igst5", "cgst25", "sgst25"),
    (12, "igst12", "cgst6", "sgst6"),
    (18, "igst18", "cgst9", "sgst9"),
    (28, "igst28", "cgst14", "sgst14"),
]


def _find_item_template_header_row(sheet) -> tuple[int, dict] | None:
    for row in range(1, min(sheet.max_row, 5) + 1):
        found = {}
        for col in range(1, sheet.max_column + 1):
            value = _normalize(sheet.cell(row, col).value)
            if not value:
                continue
            for field_name, variants in _ITEM_TEMPLATE_HEADERS.items():
                if field_name not in found and value in variants:
                    found[field_name] = col
        # This template is only usable once we can identify the invoice
        # number, party, and a per-line amount column.
        if "invoice_number" in found and "party" in found and "amount" in found:
            return row, found
    return None


def _looks_like_item_template(sheet) -> bool:
    return _find_item_template_header_row(sheet) is not None


def _parse_item_template_excel(sheet) -> list[ParsedBillRow]:
    header_row, cols = _find_item_template_header_row(sheet)

    rate_cols = {}
    for col in range(1, sheet.max_column + 1):
        rate_cols[_normalize(sheet.cell(header_row, col).value)] = col

    def get(row, name):
        col = cols.get(name)
        return sheet.cell(row, col).value if col else None

    raw_rows: list[ParsedBillRow] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        amount = _number(get(row, "amount"))
        invoice_number = str(get(row, "invoice_number") or "").strip() or None
        party = str(get(row, "party") or "").strip() or None
        stock_item = str(get(row, "stock_item") or "").strip() or None

        # A blank continuation row (e.g. a spacer, or a totals row with no
        # invoice number / amount) isn't a bill line — skip it rather than
        # creating a bogus zero-value transaction.
        if amount is None and not invoice_number:
            continue
        if amount is None:
            amount = 0.0

        date_str = _parse_date_cell(get(row, "date"))
        base_warnings = []
        if not date_str:
            base_warnings.append("No date found on this row — fill in manually before approving.")
        if not invoice_number:
            base_warnings.append("No voucher number found on this row.")

        cgst = sgst = igst = 0.0
        rate = 0.0
        for slab_rate, igst_key, cgst_key, sgst_key in _ITEM_TEMPLATE_RATE_SLABS:
            v_igst = _number(sheet.cell(row, rate_cols[igst_key]).value) if igst_key in rate_cols else None
            v_cgst = _number(sheet.cell(row, rate_cols[cgst_key]).value) if cgst_key in rate_cols else None
            v_sgst = _number(sheet.cell(row, rate_cols[sgst_key]).value) if sgst_key in rate_cols else None
            if (v_igst and abs(v_igst) > 0.004) or (v_cgst and abs(v_cgst) > 0.004) or (v_sgst and abs(v_sgst) > 0.004):
                igst = round((v_igst or 0.0), 2)
                cgst = round((v_cgst or 0.0), 2)
                sgst = round((v_sgst or 0.0), 2)
                rate = slab_rate
                break

        if rate == 0.0:
            explicit_rate = _number(get(row, "gst_rate"))
            if explicit_rate:
                rate = explicit_rate * 100 if 0 < explicit_rate <= 1 else explicit_rate

        round_off = _number(get(row, "round_off")) or 0.0
        line_total = round(amount + cgst + sgst + igst + round_off, 2)

        qty = _number(get(row, "quantity"))
        price = _number(get(row, "rate"))
        vch_type_val = str(get(row, "vch_type") or "").strip() or None

        item_line = []
        if stock_item:
            item_line = [{
                "name": stock_item,
                "hsn": str(get(row, "hsn") or "").strip() or None,
                "qty": qty,
                "unit": str(get(row, "unit") or "").strip() or None,
                "price": price,
                "amount": round(amount, 2),
                "rate": rate,
            }]

        raw_rows.append(ParsedBillRow(
            party=party, date=date_str, invoice_number=invoice_number,
            taxable_value=round(amount, 2), gst_rate=rate,
            cgst=cgst, sgst=sgst, igst=igst,
            total_value=line_total, warnings=base_warnings,
            items=item_line, vch_type=vch_type_val,
        ))

    if not raw_rows:
        raise ValueError("Found the Tally voucher template header but no item/bill rows under it.")

    return _consolidate_invoice_rows(raw_rows)


def parse_bill_excel(file_path: str) -> list[ParsedBillRow]:
    sheet = _load_first_sheet(file_path)

    if _looks_like_item_template(sheet):
        return _parse_item_template_excel(sheet)

    header = _find_header_row(sheet)
    if not header:
        raise ValueError(
            "Could not find a recognizable bill table in this Excel — need at least a party/date "
            "column plus a taxable or total amount column."
        )
    header_row, cols = header
    rate_buckets = _find_rate_buckets(sheet, header_row)

    def get(row, name):
        col = cols.get(name)
        return sheet.cell(row, col).value if col else None

    def bucket_amount(row, col):
        return _number(sheet.cell(row, col).value) if col else 0.0

    raw_rows = []
    for row in range(header_row + 1, sheet.max_row + 1):
        taxable = _number(get(row, "taxable_value"))
        total = _number(get(row, "total_value"))
        if taxable is None and total is None:
            continue

        date_str = _parse_date_cell(get(row, "date"))

        party = str(get(row, "party") or "").strip() or None
        invoice_number = str(get(row, "invoice_number") or "").strip() or None
        base_warnings = []
        if not date_str:
            base_warnings.append("No date found on this row — fill in manually before approving.")

        active_buckets = []
        for b in rate_buckets:
            bval = bucket_amount(row, b["value_col"]) or 0.0
            bcgst = bucket_amount(row, b["cgst_col"]) or 0.0
            bsgst = bucket_amount(row, b["sgst_col"]) or 0.0
            bigst = bucket_amount(row, b["igst_col"]) or 0.0
            if abs(bval) > 0.004 or abs(bcgst) > 0.004 or abs(bsgst) > 0.004 or abs(bigst) > 0.004:
                active_buckets.append((b["rate"], bval, bcgst, bsgst, bigst))

        if len(active_buckets) > 1:
            for rate, bval, bcgst, bsgst, bigst in active_buckets:
                line_total = round(bval + bcgst + bsgst + bigst, 2)
                raw_rows.append(ParsedBillRow(
                    party=party, date=date_str, invoice_number=invoice_number,
                    taxable_value=round(bval, 2), gst_rate=rate,
                    cgst=round(bcgst, 2), sgst=round(bsgst, 2), igst=round(bigst, 2),
                    total_value=line_total,
                    warnings=list(base_warnings) + [
                        f"Split from a multi-rate invoice ({len(active_buckets)} slabs) — verify against the source row."
                    ],
                    split_from_multi_rate=True,
                ))
            continue

        if len(active_buckets) == 1:
            rate, taxable, cgst, sgst, igst = active_buckets[0]
        else:
            cgst = _number(get(row, "cgst_amount")) or 0.0
            sgst = _number(get(row, "sgst_amount")) or 0.0
            igst = _number(get(row, "igst_amount")) or 0.0

            rate_cell = get(row, "gst_rate")
            if rate_cell is not None:
                rate = _number(rate_cell) or 0.0
                if 0 < rate <= 1:
                    rate *= 100
            elif taxable:
                rate = _closest_rate(((cgst + sgst + igst) / taxable) * 100)
            else:
                rate = 0.0

            if taxable is None:
                taxable = round((total or 0) - cgst - sgst - igst, 2)

        if total is None:
            total = round(taxable + cgst + sgst + igst, 2)

        raw_rows.append(ParsedBillRow(
            party=party, date=date_str, invoice_number=invoice_number,
            taxable_value=round(taxable, 2), gst_rate=rate,
            cgst=round(cgst, 2), sgst=round(sgst, 2), igst=round(igst, 2),
            total_value=round(total, 2), warnings=base_warnings,
        ))

    if not raw_rows:
        raise ValueError("Found a header row but no bill rows with a usable amount under it.")

    return _consolidate_invoice_rows(raw_rows)