"""Parse purchase invoices from the official GSTR-2B B2B Excel sheet.

Purchase values are never estimated from a bill photo.  The B2B sheet is the
source of truth for supplier invoice/tax values.
"""
import re
from dataclasses import dataclass, field
from datetime import datetime
import openpyxl
from dateutil import parser as dateparser


def _normalize(text) -> str:
    if text is None:
        return ""
    return re.sub(r"[₹()%\s\-/]", "", str(text)).lower()


def _merged_grid(sheet):
    grid = {}
    for r in sheet.merged_cells.ranges:
        value = sheet.cell(r.min_row, r.min_col).value
        for rr in range(r.min_row, r.max_row + 1):
            for cc in range(r.min_col, r.max_col + 1):
                grid[(rr, cc)] = value
    return grid


def _cell(sheet, grid, row, col):
    return grid.get((row, col), sheet.cell(row, col).value)


def _number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[₹,\s]", "", str(value))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def _date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    # GSTR-2B exports vary (2-digit years, "1 Apr 2026", Excel serial dates
    # read as plain numbers, etc.) — fall back to a fuzzy parser before
    # giving up entirely. dayfirst=True since these are Indian invoices.
    try:
        return dateparser.parse(text, dayfirst=True, fuzzy=True).date().isoformat()
    except (ValueError, OverflowError, TypeError):
        return None


@dataclass
class GstrPurchaseRow:
    supplier_gstin: str
    supplier_name: str
    invoice_number: str
    invoice_date: str | None
    invoice_value: float
    taxable_value: float
    gst_rate: float
    integrated_tax: float
    central_tax: float
    state_tax: float
    cess: float
    place_of_supply: str | None
    source_sheet: str
    warnings: list = field(default_factory=list)


HEADER_VARIANTS = {
    "supplier_gstin": ["gstinofsupplier"],
    "supplier_name": ["tradelegalname", "legalname"],
    "invoice_number": ["invoicenumber", "invoiceno"],
    "invoice_date": ["invoicedate"],
    "invoice_value": ["invoicevalue", "invoicevaluer"],
    "gst_rate": ["rate", "ratepercent"],
    "taxable_value": ["taxablevalue", "taxablevaluer"],
    "integrated_tax": ["integratedtax", "integratedtaxr"],
    "central_tax": ["centraltax", "centraltaxr"],
    "state_tax": ["stateuttax", "stateuttaxr", "statetax"],
    "cess": ["cess", "cessr"],
    "place_of_supply": ["placeofsupply"],
}


def _find_header(sheet, grid):
    for row in range(1, min(sheet.max_row, 20) + 1):
        found = {}
        for col in range(1, sheet.max_column + 1):
            value = _normalize(_cell(sheet, grid, row, col))
            if not value:
                continue
            for field, variants in HEADER_VARIANTS.items():
                if field not in found and value in variants:
                    found[field] = col
        if {"supplier_gstin", "invoice_number", "invoice_date"}.issubset(found):
            return row, found
    return None


def parse_gstr2b_purchase_excel(file_path: str) -> list[GstrPurchaseRow]:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    results = []
    sheets_found = 0

    for name in wb.sheetnames:
        norm = _normalize(name)
        # Do not read amendment/note sheets here.  They are handled by the
        # existing discount parser.
        if "b2b" not in norm or "cdnr" in norm or "cdnra" in norm:
            continue

        sheet = wb[name]
        grid = _merged_grid(sheet)
        header = _find_header(sheet, grid)
        if not header:
            continue
        sheets_found += 1
        header_row, cols = header

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            gstin = _cell(sheet, grid, row_idx, cols["supplier_gstin"])
            invoice_no = _cell(sheet, grid, row_idx, cols["invoice_number"])
            if not gstin or not str(gstin).strip():
                continue
            if not invoice_no or not str(invoice_no).strip():
                continue

            def get(field):
                col = cols.get(field)
                return _cell(sheet, grid, row_idx, col) if col else None

            item = GstrPurchaseRow(
                supplier_gstin=str(gstin).strip(),
                supplier_name=str(get("supplier_name") or "Unknown Supplier").strip(),
                invoice_number=str(invoice_no).strip(),
                invoice_date=_date(get("invoice_date")),
                invoice_value=_number(get("invoice_value")),
                taxable_value=_number(get("taxable_value")),
                gst_rate=_number(get("gst_rate")),
                integrated_tax=_number(get("integrated_tax")),
                central_tax=_number(get("central_tax")),
                state_tax=_number(get("state_tax")),
                cess=_number(get("cess")),
                place_of_supply=str(get("place_of_supply") or "").strip() or None,
                source_sheet=name,
            )

            if not item.invoice_date:
                item.warnings.append("Could not parse invoice date")
            if item.invoice_value == 0 and item.taxable_value == 0:
                item.warnings.append("Invoice value and taxable value are both zero — check manually")

            total_tax = item.integrated_tax + item.central_tax + item.state_tax
            if item.gst_rate == 0 and total_tax > 0 and item.taxable_value > 0:
                item.gst_rate = round((total_tax / item.taxable_value) * 100, 2)
                item.warnings.append(f"Rate column missing/zero — derived {item.gst_rate}% from tax amounts")

            expected = round(item.taxable_value + total_tax + item.cess, 2)
            if item.invoice_value and abs(expected - item.invoice_value) > 1.0:
                item.warnings.append(
                    f"Invoice Value (₹{item.invoice_value}) differs from taxable + tax (₹{expected}) — verify manually"
                )
            results.append(item)

    if sheets_found == 0:
        raise ValueError(
            "No B2B invoice sheet was found. Upload the actual GSTR-2B Excel export "
            "from the GST portal; purchase invoices must be in the B2B sheet."
        )
    return results
