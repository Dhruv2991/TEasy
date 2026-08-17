"""
Parses the official GSTR-2B Excel file downloaded from the GST portal
(Services > Returns > Returns Dashboard > GSTR-2B) and extracts the
Credit/Debit Note rows from the "B2B-CDNR" sheet (and "B2B-CDNRA" for
amendments, if present).

No OCR involved — this is a structured government-generated Excel file, so
we read it directly with openpyxl/pandas. This is inherently far more
reliable than photographing a screen (as in the very first uploaded image of
this sheet) — always prefer uploading the actual .xlsx file GST portal
generates over a photo of it on screen.

The GST portal's official column names for this sheet (per GSTR-2B/2A
documentation) are matched by NAME, not position, so this keeps working even
if the portal reorders/adds columns in a future export format:

    GSTIN of supplier, Trade/Legal name, Note number, Note type,
    Note Supply type, Note date, Note Value(₹) / Note Value,
    Rate(%), Taxable Value (₹) / Taxable Value,
    Integrated Tax(₹) / Integrated Tax, Central Tax(₹) / Central Tax,
    State/UT Tax(₹) / State/UT tax, Cess(₹) / Cess,
    Place of supply, Supply Attract Reverse Charge, ITC Availability

The portal's real export also has several merged title/heading rows above
the actual column headers (visible in the screenshot: rows 1-6 are titles
before the real header row) — so we scan for the header row instead of
assuming a fixed row number.
"""
import re
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl


# Sheet names to look for, in priority order. GST portal sometimes names
# these "B2B-CDNR"/"B2B - CDNR"/"B2B_CDNR" depending on export version, so we
# match loosely (case/space/punctuation-insensitive).
TARGET_SHEET_PATTERNS = ["b2bcdnr", "b2bcdnra"]

# Maps our internal field name -> list of acceptable header text variants
# (normalized: lowercased, stripped of currency symbols/spaces/punctuation)
# seen across different GSTR-2B export versions.
HEADER_MAP = {
    "supplier_gstin": ["gstinofsupplier"],
    "supplier_name": ["tradelegalname", "legalname"],
    "note_number": ["notenumber"],
    "note_type": ["notetype"],
    "note_supply_type": ["notesupplytype"],
    "note_date": ["notedate"],
    "note_value": ["notevalue", "notevaluer"],
    "gst_rate": ["rate", "rate"],
    "taxable_value": ["taxablevalue", "taxablevaluer"],
    "integrated_tax": ["integratedtax", "integratedtaxr"],
    "central_tax": ["centraltax", "centraltaxr"],
    "state_tax": ["stateuttax", "stateuttaxr", "statetax"],
    "cess": ["cess", "cessr"],
    "place_of_supply": ["placeofsupply"],
}


def _normalize(text) -> str:
    if text is None:
        return ""
    text = str(text)
    text = re.sub(r"[₹()%\s\-/]", "", text)
    return text.lower()


def _find_target_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        norm = _normalize(name)
        for pattern in TARGET_SHEET_PATTERNS:
            if pattern in norm:
                return name
    return None


def _build_merged_value_grid(sheet) -> dict[tuple[int, int], object]:
    """
    openpyxl only stores a value in the top-left cell of a merged range; all
    other cells in that range read as None. The real GSTR-2B export merges
    header cells like "GSTIN of supplier" down/across multiple header rows,
    which would otherwise break header detection. This fills every cell in
    each merged range with the range's actual value, for a subset of rows
    likely to contain headers (cheap enough to just do the whole sheet for
    typical GSTR-2B file sizes).
    """
    grid: dict[tuple[int, int], object] = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left_value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                grid[(row_idx, col_idx)] = top_left_value
    return grid


def _cell_value(sheet, merged_grid: dict, row_idx: int, col_idx: int):
    if (row_idx, col_idx) in merged_grid:
        return merged_grid[(row_idx, col_idx)]
    return sheet.cell(row=row_idx, column=col_idx).value


def _find_header_row(sheet, merged_grid: dict) -> tuple[int, dict[str, int]] | None:
    """
    Scans the first ~15 rows for the one containing recognizable column
    headers (the GST portal export has several title/merged rows above the
    real header, and some header cells like "GSTIN of supplier" are merged
    down across two header rows). Returns
    (header_row_index, {field_name: column_index}).
    """
    for row_idx in range(1, min(sheet.max_row, 15) + 1):
        col_map: dict[str, int] = {}
        for col_idx in range(1, sheet.max_column + 1):
            cell_val = _normalize(_cell_value(sheet, merged_grid, row_idx, col_idx))
            if not cell_val:
                continue
            for field_name, variants in HEADER_MAP.items():
                if field_name in col_map:
                    continue
                if cell_val in variants:
                    col_map[field_name] = col_idx

        # Require the core identifying columns to be present before we trust
        # this is really the header row (avoids matching a stray cell in a
        # title row).
        required = {"supplier_gstin", "note_number", "note_date"}
        if required.issubset(col_map.keys()):
            return row_idx, col_map

    return None


def _parse_number(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = re.sub(r"[₹,\s]", "", str(val))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    text = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


@dataclass
class GstrNoteRow:
    supplier_gstin: str
    supplier_name: str
    note_number: str
    note_type: str  # "Credit Note" or "Debit Note", as printed by GST portal
    note_date: str | None
    note_value: float
    taxable_value: float
    gst_rate: float
    integrated_tax: float
    central_tax: float
    state_tax: float
    cess: float
    place_of_supply: str | None
    source_sheet: str
    warnings: list = field(default_factory=list)


def parse_gstr2b_excel(file_path: str) -> list[GstrNoteRow]:
    """
    Returns one GstrNoteRow per credit/debit note found across the
    B2B-CDNR / B2B-CDNRA sheets. Raises ValueError with a clear message if
    the file doesn't look like a GSTR-2B export at all (wrong file
    uploaded), so the caller can surface that to the user directly rather
    than silently returning zero rows.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)

    results: list[GstrNoteRow] = []
    sheets_found = 0

    for name in wb.sheetnames:
        norm = _normalize(name)
        if not any(p in norm for p in TARGET_SHEET_PATTERNS):
            continue
        sheet = wb[name]
        merged_grid = _build_merged_value_grid(sheet)
        header = _find_header_row(sheet, merged_grid)
        if not header:
            continue
        sheets_found += 1
        header_row_idx, col_map = header

        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            gstin_cell = _cell_value(sheet, merged_grid, row_idx, col_map["supplier_gstin"])
            if not gstin_cell or not str(gstin_cell).strip():
                continue  # blank row = end of data (or a stray gap)

            def get(field_name):
                col = col_map.get(field_name)
                return _cell_value(sheet, merged_grid, row_idx, col) if col else None

            row = GstrNoteRow(
                supplier_gstin=str(gstin_cell).strip(),
                supplier_name=str(get("supplier_name") or "").strip() or "Unknown Supplier",
                note_number=str(get("note_number") or "").strip(),
                note_type=str(get("note_type") or "").strip() or "Credit Note",
                note_date=_parse_date(get("note_date")),
                note_value=_parse_number(get("note_value")),
                taxable_value=_parse_number(get("taxable_value")),
                gst_rate=_parse_number(get("gst_rate")),
                integrated_tax=_parse_number(get("integrated_tax")),
                central_tax=_parse_number(get("central_tax")),
                state_tax=_parse_number(get("state_tax")),
                cess=_parse_number(get("cess")),
                place_of_supply=str(get("place_of_supply") or "").strip() or None,
                source_sheet=name,
            )
            if not row.note_date:
                row.warnings.append("Could not parse note date")
            if row.note_value == 0 and row.taxable_value == 0:
                row.warnings.append("Note value and taxable value both zero — check manually")

            # gst_rate should come from the sheet's own "Rate(%)" column, but
            # that column isn't always present/positioned consistently across
            # GSTR-2B export versions. Since Tally needs the correct rate to
            # pick the right GST ledger, derive it from the actual tax
            # amounts as a robust fallback whenever the sheet's rate is
            # missing/zero but real tax was charged.
            total_tax = row.integrated_tax + row.central_tax + row.state_tax
            if row.gst_rate == 0 and total_tax > 0 and row.taxable_value > 0:
                row.gst_rate = round((total_tax / row.taxable_value) * 100, 2)
                row.warnings.append(f"Rate(%) column missing/zero — derived {row.gst_rate}% from tax amounts")

            # Reconciliation: note_value should equal taxable_value + all tax
            # components. Flag (don't silently "fix") any mismatch so it gets
            # a manual look, same as the photographed-bill pipelines.
            expected_total = round(row.taxable_value + total_tax + row.cess, 2)
            if row.note_value and abs(expected_total - row.note_value) > 1.0:  # >₹1 tolerance for rounding
                row.warnings.append(
                    f"Note Value (₹{row.note_value}) doesn't match taxable + tax (₹{expected_total}) — verify manually"
                )

            results.append(row)

    if sheets_found == 0:
        raise ValueError(
            "No B2B-CDNR (or B2B-CDNRA) sheet found in this file. "
            "Make sure you uploaded the actual GSTR-2B .xlsx export from the "
            "GST portal (Services > Returns > Returns Dashboard > GSTR-2B > "
            "Download Excel), not a screenshot or a different report."
        )

    return results
