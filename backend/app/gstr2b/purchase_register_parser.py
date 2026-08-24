"""Parse the shop's own periodic purchase register export.

This is NOT the GSTR-2B government file. It's whatever the shop's own
billing/accounting software (Marg, Busy, a custom POS, etc.) exports for a
date range — one row per purchase invoice, with the invoice's taxable value
already broken down *by GST rate* in separate columns, e.g.:

    Value@0%  Value@5%  CGST@2.5%  SGST@2.5%  IGST@5%
    Value@12% CGST@6%   SGST@6%    IGST@12%
    Value@18% CGST@9%   SGST@9%    IGST@18%
    Value@28% CGST@14%  SGST@14%   IGST@28%

GSTR-2B's B2B sheet only ever gives one taxable-value/tax total per invoice,
so when a supplier billed several products at different GST rates on one
invoice, GSTR-2B's effective rate doesn't land on a real slab and the
purchase parser (see purchase_parser.py) correctly flags it as uncertain.
This register is the shop's own record of the same invoice and — because it
already carries the per-rate split — is exactly what's needed to resolve
that uncertainty. See purchase_register_match.py for how the two are
reconciled.

Matching key back to GSTR-2B is the invoice number (this register calls it
"Ref No" in the common export this was built against; other exports may
call it "Invoice No" — both are matched by header alias). The reconciliation
step in purchase_register_match.py double-checks totals independently, so a
same-named-but-different column layout won't silently produce a wrong
match.
"""
import os
import re
from dataclasses import dataclass, field
from datetime import datetime, date

import openpyxl

try:
    import xlrd
except ImportError:  # pragma: no cover - xlrd is in requirements.txt
    xlrd = None


def _normalize(text) -> str:
    if text is None:
        return ""
    return re.sub(r"[₹()\s\-_/]", "", str(text)).lower()


def _number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[₹,\s]", "", str(value))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


def _excel_serial_to_date(value) -> str | None:
    """Handles the classic xlrd/Excel-serial-number date encoding used by
    a lot of older Indian accounting-software exports (.xls files store
    dates as a plain number of days since 1899-12-30, not as real dates)."""
    try:
        n = float(value)
    except (TypeError, ValueError):
        return None
    if n <= 0 or n > 80000:  # sanity bounds, ~ up to year 2119
        return None
    try:
        base = date(1899, 12, 30)
        return (base + __import__("datetime").timedelta(days=int(n))).isoformat()
    except (OverflowError, ValueError):
        return None


def _date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return _excel_serial_to_date(value)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y", "%Y-%m-%d", "%d-%b-%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    try:
        from dateutil import parser as dateparser
        return dateparser.parse(text, dayfirst=True, fuzzy=True).date().isoformat()
    except Exception:
        return None


HEADER_VARIANTS = {
    "supplier_gstin": ["gstno", "gstin", "suppliergstin", "gstinofsupplier"],
    "supplier_name": ["vendorname", "suppliername", "partyname"],
    "invoice_number": ["refno", "invoiceno", "invoicenumber", "billno"],
    "invoice_date": ["refdate", "invoicedate", "billdate"],
    "taxable_value": ["taxablevalue", "taxablevaluer"],
    "total_value": ["total", "invoicevalue"],
}

# A "slab-start" column: the taxable value billed at a given rate.
VALUE_AT_RATE_RE = re.compile(r"^value@?(\d+(?:\.\d+)?)%?$")
# Tax columns that belong to whichever slab column preceded them.
CGST_RE = re.compile(r"^cgst@?(\d+(?:\.\d+)?)%?$")
SGST_RE = re.compile(r"^sgst(?:ut)?@?(\d+(?:\.\d+)?)%?$")
IGST_RE = re.compile(r"^igst@?(\d+(?:\.\d+)?)%?$")


def _read_rows_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    for row in sheet.iter_rows(values_only=True):
        yield list(row)


def _read_rows_xls(path):
    if xlrd is None:
        raise ValueError("xlrd is required to read legacy .xls files but is not installed.")
    wb = xlrd.open_workbook(path)
    sheet = wb.sheet_by_index(0)
    for r in range(sheet.nrows):
        yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]


def _read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return list(_read_rows_xls(path))
    return list(_read_rows_xlsx(path))


def _find_header(rows):
    """Scans the first ~15 rows for the real column-header row. This
    register format typically has a couple of title/company-name rows and
    a merged "GST@5% / GST@12% / ..." banner row above the real header, so
    we score each row rather than assuming a fixed position."""
    best_idx, best_cols, best_score = None, {}, 0
    for idx in range(min(len(rows), 15)):
        row = rows[idx]
        found = {}
        rate_slabs = []  # list of (col_index, rate)
        rate_taxes = []  # list of (col_index, kind, rate)
        for c, raw in enumerate(row):
            norm = _normalize(raw)
            if not norm:
                continue
            for field_name, variants in HEADER_VARIANTS.items():
                if field_name not in found and norm in variants:
                    found[field_name] = c
            m = VALUE_AT_RATE_RE.match(norm)
            if m:
                rate_slabs.append((c, float(m.group(1))))
                continue
            m = CGST_RE.match(norm)
            if m:
                rate_taxes.append((c, "cgst", float(m.group(1))))
                continue
            m = SGST_RE.match(norm)
            if m:
                rate_taxes.append((c, "sgst", float(m.group(1))))
                continue
            m = IGST_RE.match(norm)
            if m:
                rate_taxes.append((c, "igst", float(m.group(1))))
                continue
        has_identity = "invoice_number" in found or "supplier_gstin" in found
        score = len(found) + (2 if rate_slabs else 0)
        if has_identity and rate_slabs and score > best_score:
            best_idx = idx
            best_cols = {"fields": found, "rate_slabs": rate_slabs, "rate_taxes": rate_taxes}
            best_score = score
    if best_idx is None:
        return None
    return best_idx, best_cols


def _group_rate_columns(rate_slabs, rate_taxes):
    """Groups every Value@X% column with the CGST/SGST/IGST columns that
    immediately follow it (up to the next Value@X% column, or the end of
    the row) — this is positional, not name-based, since a slab's real GST
    rate is the *sum* of CGST+SGST (or the IGST rate alone), not the half
    rate printed in "CGST@2.5%"."""
    slabs_sorted = sorted(rate_slabs, key=lambda t: t[0])
    groups = []  # list of dicts: {rate, value_col, cgst_col, sgst_col, igst_col}
    for i, (col, value_rate) in enumerate(slabs_sorted):
        next_col = slabs_sorted[i + 1][0] if i + 1 < len(slabs_sorted) else float("inf")
        group = {"rate": value_rate, "value_col": col, "cgst_col": None, "sgst_col": None, "igst_col": None}
        for tcol, kind, _trate in rate_taxes:
            if col < tcol < next_col:
                group[f"{kind}_col"] = tcol
        groups.append(group)
    return groups


@dataclass
class RegisterInvoice:
    supplier_gstin: str | None
    supplier_name: str | None
    invoice_number: str
    invoice_date: str | None
    taxable_value: float
    total_value: float
    by_rate: dict = field(default_factory=dict)  # {rate: {"taxable_value","cgst","sgst","igst"}}
    warnings: list = field(default_factory=list)
    row_number: int = 0


def parse_purchase_register(file_path: str) -> list[RegisterInvoice]:
    rows = _read_rows(file_path)
    header = _find_header(rows)
    if not header:
        raise ValueError(
            "Could not find a recognizable purchase register table in this file — expected columns "
            "like 'Ref No'/'GST No' plus a per-rate breakup (Value@5%, CGST@2.5%, SGST@2.5%, IGST@5%, "
            "Value@12%, ...). If your software exports a differently named breakup, this reader may "
            "need a small update to recognize it."
        )
    header_idx, cols = header
    fields = cols["fields"]
    rate_groups = _group_rate_columns(cols["rate_slabs"], cols["rate_taxes"])
    if not rate_groups:
        raise ValueError("Found the register's header row but no per-rate breakup columns under it.")

    def get(row, field_name):
        c = fields.get(field_name)
        if c is None or c >= len(row):
            return None
        return row[c]

    results = []
    for r in range(header_idx + 1, len(rows)):
        row = rows[r]
        if not row or all(v in (None, "") for v in row):
            continue
        invoice_no = get(row, "invoice_number")
        if not invoice_no or not str(invoice_no).strip():
            continue  # spacer/total row

        by_rate = {}
        for g in rate_groups:
            def cell(colkey):
                c = g[colkey]
                if c is None or c >= len(row):
                    return 0.0
                return _number(row[c])

            taxable = cell("value_col")
            cgst = cell("cgst_col")
            sgst = cell("sgst_col")
            igst = cell("igst_col")
            if taxable == 0 and cgst == 0 and sgst == 0 and igst == 0:
                continue
            by_rate[g["rate"]] = {
                "taxable_value": round(taxable, 2),
                "cgst": round(cgst, 2),
                "sgst": round(sgst, 2),
                "igst": round(igst, 2),
            }

        item = RegisterInvoice(
            supplier_gstin=str(get(row, "supplier_gstin") or "").strip() or None,
            supplier_name=str(get(row, "supplier_name") or "").strip() or None,
            invoice_number=str(invoice_no).strip(),
            invoice_date=_date(get(row, "invoice_date")),
            taxable_value=_number(get(row, "taxable_value")),
            total_value=_number(get(row, "total_value")),
            by_rate=by_rate,
            row_number=r + 1,
        )
        if not item.by_rate:
            item.warnings.append("No non-zero rate columns found for this row — skipped from matching")
        results.append(item)

    if not results:
        raise ValueError(
            "Found a header row but no invoice rows under it with a recognizable invoice number."
        )
    return results
