"""Parse a supplier's own invoice Excel to recover the real per-line-item
GST-rate breakdown that GSTR-2B's B2B sheet structurally cannot provide
(GSTR-2B is invoice-level, not line-item-level).

Supplier files vary wildly — different column names, column order, extra
merchandising columns, sometimes a totals row, sometimes not. This parser
does NOT assume a fixed template. It scores each row against a set of
known header aliases, picks the best-matching header row, and then reads
line items by GST rate, exactly as printed on the invoice (5%, 12%, 18%...).

This module only *extracts* what's on the sheet. It makes no claim that
the extraction is correct for Tally — that's the job of
`supplier_match.reconcile_with_transaction`, which cross-checks the
extracted totals against the GSTR-2B transaction already on file before
anything is allowed to be pushed.
"""
import re
from dataclasses import dataclass, field

import openpyxl


def _normalize(text) -> str:
    if text is None:
        return ""
    return re.sub(r"[₹()%\s\-_/.]", "", str(text)).lower()


def _number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[₹,\s]", "", str(value))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


# Column aliases seen across common supplier invoice exports (Tally, Busy,
# Zoho, Vyapar, plain Excel templates). Extend this list as new supplier
# formats show up — it costs nothing and never breaks older matches.
LINE_HEADER_VARIANTS = {
    "description": ["itemdescription", "description", "particulars", "itemname", "product"],
    "taxable_value": ["taxablevalue", "taxableamount", "amount", "netamount", "value"],
    "gst_rate": ["gstrate", "taxrate", "rate", "gst%", "igstrate", "cgstrate+sgstrate"],
    "cgst_amount": ["cgstamount", "cgst"],
    "sgst_amount": ["sgstamount", "sgst", "sgstutamount"],
    "igst_amount": ["igstamount", "igst"],
    "invoice_number": ["invoiceno", "invoicenumber", "billno", "billnumber"],
    "invoice_date": ["invoicedate", "billdate", "date"],
    "supplier_gstin": ["gstin", "suppliergstin", "gstno"],
}

# Real GST slabs, same list used for reconciling GSTR-2B rows.
STANDARD_GST_RATES = [0, 0.25, 1, 1.5, 3, 5, 6, 7.5, 12, 13.8, 18, 28]


def _closest_rate(pct: float) -> float | None:
    if pct is None:
        return None
    nearest = min(STANDARD_GST_RATES, key=lambda r: abs(r - pct))
    return nearest if abs(nearest - pct) <= 0.5 else None


def _find_header_row(sheet):
    best_row, best_cols, best_score = None, {}, 0
    for row in range(1, min(sheet.max_row, 25) + 1):
        found = {}
        for col in range(1, sheet.max_column + 1):
            value = _normalize(sheet.cell(row, col).value)
            if not value:
                continue
            for field_name, variants in LINE_HEADER_VARIANTS.items():
                if field_name not in found and value in variants:
                    found[field_name] = col
        # A usable header row needs at minimum a taxable/amount column and
        # something that tells us the rate (either a rate column, or split
        # CGST/SGST/IGST amount columns to derive it from).
        has_amount = "taxable_value" in found
        has_rate_signal = "gst_rate" in found or "cgst_amount" in found or "igst_amount" in found
        score = len(found)
        if has_amount and has_rate_signal and score > best_score:
            best_row, best_cols, best_score = row, found, score
    if best_row is None:
        return None
    return best_row, best_cols


@dataclass
class SupplierInvoiceLine:
    description: str
    taxable_value: float
    gst_rate: float
    cgst: float
    sgst: float
    igst: float


@dataclass
class SupplierInvoice:
    invoice_number: str | None
    invoice_date: str | None
    supplier_gstin: str | None
    lines: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    @property
    def by_rate(self) -> dict:
        """Line items grouped and summed by GST rate — this is the shape the
        voucher builder needs (one Tally ledger set per distinct rate)."""
        grouped: dict[float, dict] = {}
        for line in self.lines:
            bucket = grouped.setdefault(line.gst_rate, {"taxable_value": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0})
            bucket["taxable_value"] += line.taxable_value
            bucket["cgst"] += line.cgst
            bucket["sgst"] += line.sgst
            bucket["igst"] += line.igst
        return grouped


def parse_supplier_invoice_excel(file_path: str) -> SupplierInvoice:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    sheet = wb[wb.sheetnames[0]]  # supplier invoices are effectively always single-sheet

    header = _find_header_row(sheet)
    if not header:
        raise ValueError(
            "Could not find a recognizable line-item table in this Excel (need at least an "
            "amount/taxable-value column plus a GST rate, or CGST/SGST/IGST amount columns)."
        )
    header_row, cols = header

    def get(sheet_row, field_name):
        col = cols.get(field_name)
        return sheet.cell(sheet_row, col).value if col else None

    # Invoice-level fields are often printed once above the line-item table,
    # not repeated per row — scan the header row's own row plus the rows
    # above it for a one-off match too, falling back to per-line values.
    invoice_number = None
    invoice_date = None
    supplier_gstin = None
    for row in range(1, header_row + 1):
        for field_name, target in (
            ("invoice_number", "invoice_number"),
            ("invoice_date", "invoice_date"),
            ("supplier_gstin", "supplier_gstin"),
        ):
            col = cols.get(field_name)
            if col:
                val = sheet.cell(row, col).value
                if val and target == "invoice_number" and not invoice_number:
                    invoice_number = str(val).strip()
                if val and target == "invoice_date" and not invoice_date:
                    invoice_date = str(val).strip()
                if val and target == "supplier_gstin" and not supplier_gstin:
                    supplier_gstin = str(val).strip()

    lines = []
    warnings = []
    for row in range(header_row + 1, sheet.max_row + 1):
        taxable = _number(get(row, "taxable_value"))
        if taxable == 0:
            continue  # blank spacer row or a totals row with no per-item amount
        desc = str(get(row, "description") or "").strip()
        if _normalize(desc) in ("total", "grandtotal", ""):
            # Totals rows often reuse the amount column — skip anything that
            # looks like a summary line rather than a real item.
            if _normalize(desc) in ("total", "grandtotal"):
                continue

        cgst = _number(get(row, "cgst_amount"))
        sgst = _number(get(row, "sgst_amount"))
        igst = _number(get(row, "igst_amount"))
        rate_cell = get(row, "gst_rate")

        if rate_cell is not None:
            rate = _number(rate_cell)
            if rate <= 1 and rate > 0:  # e.g. "0.18" instead of "18"
                rate *= 100
        else:
            # No explicit rate column — derive it from the CGST+SGST+IGST vs
            # taxable value, same logic as the GSTR-2B fallback.
            effective = ((cgst + sgst + igst) / taxable) * 100 if taxable else 0
            rate = _closest_rate(effective)
            if rate is None:
                warnings.append(f"Row {row} ('{desc or 'unnamed item'}'): could not resolve a clean GST rate — skipped")
                continue

        lines.append(SupplierInvoiceLine(
            description=desc or "Item",
            taxable_value=taxable,
            gst_rate=rate,
            cgst=cgst,
            sgst=sgst,
            igst=igst,
        ))

    if not lines:
        raise ValueError("Found a header row but no line items with a non-zero taxable value under it.")

    return SupplierInvoice(
        invoice_number=invoice_number,
        invoice_date=invoice_date,
        supplier_gstin=supplier_gstin,
        lines=lines,
        warnings=warnings,
    )
