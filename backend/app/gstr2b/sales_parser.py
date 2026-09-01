"""Parse sales invoices from simple sales registers AND Tally Sales Excel exports.

Handles case/spacing-insensitive header matching, asterisk stripping,
dynamic multi-slab tax column aggregation (e.g. CGST 9%, IGST 18%), and
derived taxable/total values.
"""
import re
from dataclasses import dataclass, field

from .purchase_parser import _normalize, _merged_grid, _cell, _number, _date


@dataclass
class SalesRow:
    party: str
    invoice_number: str
    invoice_date: str | None
    total_value: float
    taxable_value: float
    gst_rate: float
    cgst: float
    sgst: float
    igst: float
    cess: float
    source_sheet: str
    warnings: list = field(default_factory=list)


HEADER_VARIANTS = {
    "party": ["party", "partyname", "customer", "customername", "buyer", "buyername", "name", "partyledger"],
    "invoice_number": ["invoicenumber", "invoiceno", "billno", "billnumber", "invno", "vchno", "voucherno"],
    "invoice_date": ["invoicedate", "date", "billdate", "vchdate", "voucherdate"],
    "taxable_value": ["taxablevalue", "taxablevaluer", "taxableamount", "taxable"],
    "gst_rate": ["rate", "ratepercent", "gstrate"],
    "cgst": ["cgst", "centraltax", "centraltaxr"],
    "sgst": ["sgst", "stateuttax", "stateuttaxr", "statetax"],
    "igst": ["igst", "integratedtax", "integratedtaxr"],
    "cess": ["cess", "cessr"],
    "total_value": ["totalvalue", "invoicevalue", "total", "billamount", "grandtotal", "amount"],
}

# Only core identifiers are strictly required; financial fields can be derived
REQUIRED_HEADERS = {"party", "invoice_number", "invoice_date"}


def _clean_header(val: str | None) -> str:
    """Strips asterisks, special characters, and extra spaces for matching."""
    if not val:
        return ""
    normalized = _normalize(val)
    return re.sub(r"[^a-z0-9%]", "", normalized)


def _find_header(sheet, grid):
    for row in range(1, min(sheet.max_row, 20) + 1):
        found = {}
        tax_slab_cols = {"cgst": [], "sgst": [], "igst": []}

        for col in range(1, sheet.max_column + 1):
            raw_val = _cell(sheet, grid, row, col)
            clean_val = _clean_header(raw_val)
            if not clean_val:
                continue

            # Check standard header variants
            for field_name, variants in HEADER_VARIANTS.items():
                if field_name not in found and any(v in clean_val for v in variants):
                    found[field_name] = col

            # Detect multi-slab tax columns (e.g., "cgst9%", "igst18%", "sgst25%")
            for tax_type in ["cgst", "sgst", "igst"]:
                if tax_type in clean_val and re.search(r"\d", clean_val):
                    tax_slab_cols[tax_type].append(col)

        # Ensure we have at least party, invoice number, date, and SOME total or taxable value indicator
        if REQUIRED_HEADERS.issubset(found) and ("total_value" in found or "taxable_value" in found):
            return row, found, tax_slab_cols
            
    return None


def parse_sales_excel(file_path: str) -> list[SalesRow]:
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    results = []
    sheets_found = 0

    for name in wb.sheetnames:
        sheet = wb[name]
        grid = _merged_grid(sheet)
        header_info = _find_header(sheet, grid)
        if not header_info:
            continue

        sheets_found += 1
        header_row, cols, tax_slab_cols = header_info

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            party = _cell(sheet, grid, row_idx, cols.get("party"))
            invoice_no = _cell(sheet, grid, row_idx, cols.get("invoice_number"))

            if not party or not str(party).strip():
                continue
            if not invoice_no or not str(invoice_no).strip():
                continue

            def get(field_name):
                col = cols.get(field_name)
                return _cell(sheet, grid, row_idx, col) if col else None

            # Aggregate taxes from single tax column OR multi-slab tax columns
            cgst_val = _number(get("cgst")) + sum(_number(_cell(sheet, grid, row_idx, c)) for c in tax_slab_cols["cgst"])
            sgst_val = _number(get("sgst")) + sum(_number(_cell(sheet, grid, row_idx, c)) for c in tax_slab_cols["sgst"])
            igst_val = _number(get("igst")) + sum(_number(_cell(sheet, grid, row_idx, c)) for c in tax_slab_cols["igst"])

            total_val = _number(get("total_value"))
            taxable_val = _number(get("taxable_value"))
            total_tax = cgst_val + sgst_val + igst_val
            cess_val = _number(get("cess"))

            warnings = []

            # Derive Taxable Value if missing (Total Value - Total Tax - Cess)
            if taxable_val == 0 and total_val > 0:
                taxable_val = round(total_val - total_tax - cess_val, 2)
                warnings.append("Taxable value derived from Total Value minus taxes")

            # Derive Total Value if missing (Taxable Value + Total Tax + Cess)
            if total_val == 0 and (taxable_val > 0 or total_tax > 0):
                total_val = round(taxable_val + total_tax + cess_val, 2)
                warnings.append("Total value derived from Taxable Value + tax")

            # Calculate GST Rate if missing
            gst_rate_val = _number(get("gst_rate"))
            if gst_rate_val == 0 and total_tax > 0 and taxable_val > 0:
                gst_rate_val = round((total_tax / taxable_val) * 100, 2)
                warnings.append(f"Derived GST rate ({gst_rate_val}%) from tax amounts")

            item = SalesRow(
                party=str(party).strip(),
                invoice_number=str(invoice_no).strip(),
                invoice_date=_date(get("invoice_date")),
                total_value=total_val,
                taxable_value=taxable_val,
                gst_rate=gst_rate_val,
                cgst=cgst_val,
                sgst=sgst_val,
                igst=igst_val,
                cess=cess_val,
                source_sheet=name,
                warnings=warnings,
            )

            if not item.invoice_date:
                item.warnings.append("Could not parse invoice date")

            # Reconciliation sanity check
            expected_total = round(item.taxable_value + total_tax + item.cess, 2)
            if item.total_value and abs(expected_total - item.total_value) > 1.5:
                item.warnings.append(
                    f"Total ({item.total_value}) differs from taxable + tax ({expected_total}) — verify manually"
                )

            results.append(item)

    if sheets_found == 0:
        raise ValueError(
            "Could not find a valid sales sheet. Expected columns for Party/Customer, Invoice/Voucher Number, and Invoice Date."
        )

    return results